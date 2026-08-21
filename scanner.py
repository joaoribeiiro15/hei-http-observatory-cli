#!/usr/bin/env python3
"""
Mozilla Observatory Scanner
Reads all CSVs from ./source/, scans each URL via the MDN HTTP Observatory API,
prints results to the terminal, and writes a consolidated CSV to ./results/.
"""

import csv
import json
import re
import sys
import time
import os
import glob
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

API_BASE = "https://observatory-api.mdn.mozilla.net/api/v2"
SCAN_COOLDOWN = 2          # seconds between requests (be polite to the API)
RESULTS_DIR = Path("results")
SOURCE_DIR  = Path("source")


# Maps API result codes → (human-readable reason, recommendation)
# Mirrors what Mozilla Observatory shows in its Test / Score / Reason / Recommendation table.
RESULT_DETAILS: dict[str, tuple[str, str]] = {
    # ── Content Security Policy ──────────────────────────────────────────────
    "csp-not-implemented":
        ("Content Security Policy (CSP) header not implemented",
         "Implement one, see MDN's Content Security Policy (CSP) documentation."),
    "csp-header-invalid":
        ("Content Security Policy (CSP) header cannot be parsed successfully",
         "Fix the syntax errors in your CSP header."),
    "csp-implemented-with-no-unsafe":
        ("Content Security Policy (CSP) implemented without 'unsafe-inline' or 'unsafe-eval'",
         "None"),
    "csp-implemented-with-unsafe-eval":
        ("CSP implemented but allows 'unsafe-eval'",
         "Remove 'unsafe-eval' from your CSP policy."),
    "csp-implemented-with-unsafe-inline":
        ("CSP implemented but allows 'unsafe-inline'",
         "Remove 'unsafe-inline' and use nonces or hashes instead."),
    "csp-implemented-with-unsafe-inline-in-style-src-only":
        ("CSP implemented with 'unsafe-inline' inside style-src only",
         "Consider removing 'unsafe-inline' from style-src."),
    "csp-implemented-with-insecure-scheme":
        ("CSP implemented but allows loading resources over HTTP",
         "Change http:// sources to https:// in your CSP."),
    "csp-implemented-with-insecure-scheme-in-passive-content-only":
        ("CSP implemented but allows passive content (images) over HTTP",
         "Change http:// image sources to https://."),

    # ── Cookies ──────────────────────────────────────────────────────────────
    "cookies-not-found":
        ("No cookies detected", "None"),
    "cookies-session-without-secure-flag":
        ("Session cookie set without using the Secure flag or set over HTTP.",
         "Use Secure flag and set up HSTS."),
    "cookies-session-without-secure-flag-but-protected-by-hsts":
        ("Session cookie set without Secure flag but protected by HSTS",
         "Add the Secure flag to session cookies."),
    "cookies-without-secure-flag-but-protected-by-hsts":
        ("Cookies set without Secure flag but protected by HSTS",
         "Add the Secure flag to all cookies."),
    "cookies-without-secure-flag":
        ("Cookies set without using the Secure flag",
         "Use Secure flag and set up HSTS."),
    "cookies-session-without-httponly-flag":
        ("Session cookie set without the HttpOnly flag",
         "Add the HttpOnly flag to session cookies."),
    "cookies-anticsrf-without-secure-flag":
        ("Anti-CSRF token cookie set without Secure flag",
         "Add the Secure flag to anti-CSRF cookies."),
    "cookies-samesite-flag-invalid":
        ("SameSite flag set to an invalid value",
         "Set SameSite to Strict, Lax, or None."),
    "cookies-secure-with-httponly-sessions":
        ("Cookies use Secure flag and session cookies use HttpOnly", "None"),
    "cookies-secure-with-httponly-sessions-and-samesite":
        ("All cookies use Secure, HttpOnly, and SameSite", "None"),

    # ── Cross-Origin Resource Sharing (CORS) ─────────────────────────────────
    "cross-origin-resource-sharing-not-implemented":
        ("Content is not visible via cross-origin resource sharing (CORS) files or headers.",
         "None"),
    "cross-origin-resource-sharing-implemented-with-public-access":
        ("Content is publicly accessible via CORS",
         "Restrict CORS to trusted origins if the content is not public."),
    "cross-origin-resource-sharing-implemented-with-restricted-access":
        ("Content is accessible to a limited set of origins via CORS", "None"),
    "cross-origin-resource-sharing-implemented-with-universal-access":
        ("Content is accessible to all origins via CORS (Access-Control-Allow-Origin: *)",
         "Restrict CORS to trusted origins on authenticated endpoints."),

    # ── Redirection ──────────────────────────────────────────────────────────
    "redirection-to-https":
        ("Redirects to HTTPS correctly", "None"),
    "redirection-not-needed-no-http":
        ("Site only accessible over HTTPS", "None"),
    "redirection-off-host-from-http":
        ("HTTP does not redirect to HTTPS on the same host before redirecting elsewhere",
         "Redirect to HTTPS on the same hostname first."),
    "redirection-missing":
        ("HTTP does not redirect to HTTPS",
         "Enable HTTPS and redirect all HTTP traffic to HTTPS."),
    "redirection-invalid-cert":
        ("Invalid certificate chain encountered during redirection.",
         "Install a valid TLS certificate on the server. Let's Encrypt is a good choice, "
         "as are certificates managed by your cloud provider or commercially sold ones."),
    "redirection-not-to-https-on-initial-redirection":
        ("HTTP redirects to another HTTP URL before reaching HTTPS",
         "Redirect directly from HTTP to HTTPS."),
    "redirection-not-to-https":
        ("HTTP does not redirect to HTTPS", "Enable HTTPS and redirect HTTP traffic."),

    # ── Referrer Policy ──────────────────────────────────────────────────────
    "referrer-policy-not-implemented":
        ("Referrer-Policy header not implemented.",
         "Set to strict-origin-when-cross-origin at a minimum."),
    "referrer-policy-header-invalid":
        ("Referrer-Policy header cannot be recognized",
         "Use a valid Referrer-Policy value."),
    "referrer-policy-unsafe":
        ("Referrer-Policy set to an unsafe value (sends full URL as referrer)",
         "Use strict-origin-when-cross-origin or stricter."),
    "referrer-policy-private":
        ("Referrer-Policy set to a value that protects privacy", "None"),
    "referrer-policy-safe":
        ("Referrer-Policy set to a safe non-default value", "None"),

    # ── Strict Transport Security (HSTS) ─────────────────────────────────────
    "hsts-not-implemented":
        ("Strict-Transport-Security header not implemented",
         "Add HSTS with max-age=63072000; includeSubDomains; preload."),
    "hsts-not-implemented-no-https":
        ("Strict-Transport-Security header not implemented — site does not support HTTPS",
         "Enable HTTPS then add HSTS."),
    "hsts-header-invalid":
        ("Strict-Transport-Security header cannot be parsed",
         "Fix the HSTS header syntax."),
    "hsts-not-implemented-header-invalid":
        ("Strict-Transport-Security header cannot be parsed",
         "Fix the HSTS header syntax."),
    "hsts-invalid-cert":
        ("Strict-Transport-Security header cannot be set, as site contains an invalid certificate chain.",
         "HSTS can only work with a valid TLS certificate on the server. Let's Encrypt is a good "
         "choice, as are certificates managed by your cloud provider or commercially sold ones."),
    "hsts-implemented-max-age-less-than-six-months":
        ("HSTS implemented but max-age is less than 6 months",
         "Set max-age to at least 15768000 (6 months)."),
    "hsts-implemented-max-age-at-least-six-months":
        ("HSTS implemented with max-age of at least six months", "None"),
    "hsts-preloaded":
        ("HSTS preloaded", "None"),

    # ── Subresource Integrity (SRI) ──────────────────────────────────────────
    "subresource-integrity-not-implemented-but-no-scripts-loaded":
        ("No scripts loaded from external origins", "None"),
    "subresource-integrity-not-implemented-but-all-scripts-loaded-from-secure-origin":
        ("Subresource Integrity (SRI) not implemented, but all scripts are loaded from a similar origin.",
         "Add SRI for bonus points."),
    "subresource-integrity-not-implemented-response-not-html":
        ("SRI check skipped — response is not HTML", "None"),
    "subresource-integrity-not-implemented":
        ("Subresource Integrity (SRI) not implemented, but all scripts are loaded from a similar origin.",
         "Add SRI for bonus points."),
    "subresource-integrity-implemented-with-all-scripts-loaded-securely":
        ("SRI implemented and all scripts loaded securely", "None"),
    "subresource-integrity-implemented-but-external-scripts-not-loaded":
        ("SRI implemented, no external scripts found", "None"),

    # ── X-Content-Type-Options ───────────────────────────────────────────────
    "x-content-type-options-not-implemented":
        ("X-Content-Type-Options header not implemented.", "Set to nosniff."),
    "x-content-type-options-header-invalid":
        ("X-Content-Type-Options header cannot be recognized",
         "Set the header value to 'nosniff'."),
    "x-content-type-options-nosniff":
        ("X-Content-Type-Options header set to 'nosniff'", "None"),

    # ── X-Frame-Options ──────────────────────────────────────────────────────
    "x-frame-options-not-implemented":
        ("X-Frame-Options (XFO) header not implemented.", "Implement frame-ancestors CSP."),
    "x-frame-options-header-invalid":
        ("X-Frame-Options header cannot be recognized", "Set to DENY or SAMEORIGIN."),
    "x-frame-options-allow-from-origin":
        ("X-Frame-Options set to ALLOW-FROM — not widely supported",
         "Use CSP frame-ancestors instead."),
    "x-frame-options-sameorigin-or-deny":
        ("X-Frame-Options set to SAMEORIGIN or DENY", "None"),
    "x-frame-options-implemented-via-csp":
        ("X-Frame-Options protection implemented via CSP frame-ancestors", "None"),

    # ── Cross-Origin-Opener-Policy (COOP) ────────────────────────────────────
    "cross-origin-opener-policy-not-implemented":
        ("Cross-Origin-Opener-Policy (COOP) header not implemented",
         "Set Cross-Origin-Opener-Policy: same-origin to isolate the browsing context."),
    "cross-origin-opener-policy-same-origin":
        ("Cross-Origin-Opener-Policy set to same-origin", "None"),
    "cross-origin-opener-policy-same-origin-allow-popups":
        ("Cross-Origin-Opener-Policy set to same-origin-allow-popups", "None"),

    # ── Cross-Origin-Resource-Policy (CORP) ──────────────────────────────────
    "cross-origin-resource-policy-not-implemented":
        ("Cross Origin Resource Policy (CORP) is not implemented (defaults to cross-origin).",
         "None"),
    "cross-origin-resource-policy-same-origin":
        ("Cross-Origin-Resource-Policy set to same-origin", "None"),
    "cross-origin-resource-policy-same-site":
        ("Cross-Origin-Resource-Policy set to same-site", "None"),
    "cross-origin-resource-policy-cross-origin":
        ("Cross-Origin-Resource-Policy set to cross-origin (explicit)", "None"),
}

# Human-readable display names for each test key
TEST_DISPLAY_NAMES = {
    "content-security-policy":        "Content Security Policy (CSP)",
    "cookies":                        "Cookies",
    "cross-origin-resource-sharing":  "Cross Origin Resource Sharing (CORS)",
    "redirection":                    "Redirection",
    "referrer-policy":                "Referrer Policy",
    "strict-transport-security":      "Strict Transport Security (HSTS)",
    "subresource-integrity":          "Subresource Integrity",
    "x-content-type-options":         "X-Content-Type-Options",
    "x-frame-options":                "X-Frame-Options",
    "cross-origin-opener-policy":     "Cross-Origin-Opener-Policy",
    "cross-origin-resource-policy":   "Cross Origin Resource Policy",
}

# Ordered list of known tests (controls CSV column order)
KNOWN_TESTS = [
    "content-security-policy",
    "cookies",
    "cross-origin-resource-sharing",
    "redirection",
    "referrer-policy",
    "strict-transport-security",
    "subresource-integrity",
    "x-content-type-options",
    "x-frame-options",
    "cross-origin-opener-policy",
    "cross-origin-resource-policy",
]

# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def write_xlsx(rows: list[dict], fieldnames: list[str], path: Path) -> None:
    from collections import Counter

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    GRADE_FILL = {
        "A+": PatternFill("solid", fgColor="C6EFCE"),
        "A":  PatternFill("solid", fgColor="C6EFCE"),
        "A-": PatternFill("solid", fgColor="C6EFCE"),
        "B+": PatternFill("solid", fgColor="BDD7EE"),
        "B":  PatternFill("solid", fgColor="BDD7EE"),
        "B-": PatternFill("solid", fgColor="BDD7EE"),
        "C+": PatternFill("solid", fgColor="FFEB9C"),
        "C":  PatternFill("solid", fgColor="FFEB9C"),
        "C-": PatternFill("solid", fgColor="FFEB9C"),
        "D+": PatternFill("solid", fgColor="FFC7CE"),
        "D":  PatternFill("solid", fgColor="FFC7CE"),
        "D-": PatternFill("solid", fgColor="FFC7CE"),
        "F":  PatternFill("solid", fgColor="FFC7CE"),
    }
    ERROR_FILL = PatternFill("solid", fgColor="EDEDED")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=13, color="1F3864")
    sub_font   = Font(bold=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center
    ws.row_dimensions[1].height = 22

    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])
        row_idx = ws.max_row
        grade = row.get("grade", "")
        error = row.get("error", "")
        fill = GRADE_FILL.get(grade) if grade else (ERROR_FILL if error else None)
        for cell in ws[row_idx]:
            cell.alignment = left
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10

    # Grade distribution table
    grade_order = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]
    grade_counts = Counter(r["grade"] for r in rows if r.get("grade"))
    present_grades = [g for g in grade_order if g in grade_counts]
    n_grades = len(present_grades)

    ws2["A1"] = "Grade Distribution"
    ws2["A1"].font = title_font
    ws2["A2"] = "Grade"
    ws2["B2"] = "Count"
    ws2["A2"].font = sub_font
    ws2["B2"].font = sub_font

    for i, grade in enumerate(present_grades, start=3):
        ws2.cell(row=i, column=1, value=grade)
        ws2.cell(row=i, column=2, value=grade_counts[grade])

    chart1 = BarChart()
    chart1.title = "Grade Distribution"
    chart1.y_axis.title = "Institutions"
    chart1.x_axis.title = "Grade"
    chart1.width = 22
    chart1.height = 14
    data1 = Reference(ws2, min_col=2, min_row=2, max_row=2 + n_grades)
    cats1 = Reference(ws2, min_col=1, min_row=3, max_row=2 + n_grades)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws2.add_chart(chart1, "D2")

    # Anchor chart2 below chart1 based on its actual height (1 cm ≈ 1.89 rows at default row height)
    chart2_anchor_row = 2 + int(chart1.height / 0.529) + 3

    # Scan status table
    status_row = n_grades + 4
    graded_count = sum(1 for r in rows if r.get("grade"))
    error_count  = sum(1 for r in rows if r.get("error"))

    ws2.cell(row=status_row, column=1, value="Scan Status").font = title_font
    ws2.cell(row=status_row + 1, column=1, value="Status").font = sub_font
    ws2.cell(row=status_row + 1, column=2, value="Count").font = sub_font
    ws2.cell(row=status_row + 2, column=1, value="Graded")
    ws2.cell(row=status_row + 2, column=2, value=graded_count)
    ws2.cell(row=status_row + 3, column=1, value="Error")
    ws2.cell(row=status_row + 3, column=2, value=error_count)

    chart2 = PieChart()
    chart2.title = "Scan Status"
    chart2.width = 14
    chart2.height = 12
    data2 = Reference(ws2, min_col=2, min_row=status_row + 1, max_row=status_row + 3)
    cats2 = Reference(ws2, min_col=1, min_row=status_row + 2, max_row=status_row + 3)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws2.add_chart(chart2, f"D{chart2_anchor_row}")

    wb.save(path)


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

class _HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self._parts: list[str] = []

    def handle_data(self, data: str) -> None:
        self._parts.append(data)

    def get_text(self) -> str:
        return re.sub(r"\s+", " ", " ".join(self._parts)).strip()


def strip_html(html_str: str) -> str:
    """Convert an HTML string to plain text, collapsing whitespace."""
    if not html_str:
        return ""
    s = _HTMLStripper()
    s.feed(html_str)
    return s.get_text()


def scan_host(hostname: str) -> dict:
    """POST to Observatory v2 to trigger a fresh scan and return the scan dict."""
    url = f"{API_BASE}/scan?host={hostname}"
    req = Request(url, method="POST", headers={"User-Agent": "observatory-scanner/1.0"})
    try:
        with urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except HTTPError as exc:
        return {"error": f"http-{exc.code}", "message": str(exc)}
    except URLError as exc:
        return {"error": "network-error", "message": str(exc.reason)}
    except Exception as exc:
        return {"error": "unknown", "message": str(exc)}


def get_analyze_data(hostname: str) -> dict:
    """GET /api/v2/analyze to retrieve per-test details with score_description,
    recommendation, title, and link fields — exactly what the Observatory website shows."""
    url = f"{API_BASE}/analyze?host={hostname}"
    req = Request(url, headers={"User-Agent": "observatory-scanner/1.0"})
    try:
        with urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode())
            return data.get("tests", {})
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# CSV discovery
# ---------------------------------------------------------------------------

def find_source_csvs() -> list[Path]:
    """Return all CSV files found in SOURCE_DIR."""
    files = sorted(SOURCE_DIR.glob("*.csv"))
    if not files:
        print(f"[WARN] No CSV files found in '{SOURCE_DIR}/'.")
    return files


def read_institutions(csv_path: Path) -> list[dict]:
    """Parse a CSV file and return a list of row dicts."""
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    # Accept 'url' or 'URL' column
    for row in rows:
        keys = list(row.keys())
        lower_map = {k.lower(): k for k in keys}
        if "url" not in row and "url" in lower_map:
            row["url"] = row[lower_map["url"]]
    return rows


# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------

RESET  = "\033[0m"
BOLD   = "\033[1m"
RED    = "\033[91m"
YELLOW = "\033[93m"
GREEN  = "\033[92m"
CYAN   = "\033[96m"
DIM    = "\033[2m"

def grade_color(grade: str) -> str:
    if not grade:
        return DIM
    g = grade[0].upper()
    if g == "A":
        return GREEN
    if g == "B":
        return CYAN
    if g in ("C", "D"):
        return YELLOW
    return RED


def print_header():
    print()
    print(f"{BOLD}{'='*70}{RESET}")
    print(f"{BOLD}  Mozilla HTTP Observatory Scanner{RESET}")
    print(f"  Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{BOLD}{'='*70}{RESET}")
    print()


def _resolve_test(tdata: dict) -> tuple[str, str]:
    """Return (reason, recommendation) from live API data, falling back to static map."""
    result_code = tdata.get("result", "")

    # Primary: use the HTML fields the API returns (same text as the Observatory website)
    api_reason = strip_html(tdata.get("score_description", ""))
    api_recomm = strip_html(tdata.get("recommendation", ""))

    if api_reason:
        reason = api_reason
    elif result_code in RESULT_DETAILS:
        reason = RESULT_DETAILS[result_code][0]
    else:
        reason = result_code.replace("-", " ").capitalize()

    if api_recomm and api_recomm.lower() not in ("none", ""):
        recomm = api_recomm
    elif result_code in RESULT_DETAILS:
        recomm = RESULT_DETAILS[result_code][1]
    else:
        recomm = "Review the Observatory report for details."

    return reason, recomm


def _print_test_table(tests: dict) -> None:
    """Print all tests in a table matching the Mozilla Observatory layout."""
    import textwrap

    W_TEST   = 34
    W_SCORE  = 8
    W_REASON = 44
    W_RECOMM = 46

    header = f"  {'Test':<{W_TEST}} {'Score':<{W_SCORE}} {'Reason':<{W_REASON}} Recommendation"
    sep    = f"  {'-'*W_TEST} {'-'*W_SCORE} {'-'*W_REASON} {'-'*W_RECOMM}"
    print()
    print(f"{BOLD}{header}{RESET}")
    print(sep)

    # Show tests in preferred order, then any extras the API returned
    ordered_keys = [k for k in KNOWN_TESTS if k in tests]
    extra_keys   = [k for k in tests if k not in KNOWN_TESTS]
    for test_key in ordered_keys + extra_keys:
        tdata = tests[test_key]

        display     = tdata.get("title") or TEST_DISPLAY_NAMES.get(test_key, test_key)
        passed      = tdata.get("pass", True)
        modifier    = tdata.get("score_modifier", None)
        mdn_link    = tdata.get("link", "")

        reason, recomm = _resolve_test(tdata)
        if passed is True and not recomm:
            recomm = "None"

        # Score column: show modifier + icon
        if modifier is None or modifier == "":
            score_str = "-"
        else:
            score_str = f"{modifier:+d}" if isinstance(modifier, int) else str(modifier)
        icon      = f"{GREEN}✓{RESET}" if passed else f"{RED}✗{RESET}"
        score_col = f"{score_str} {icon}"
        sc_plain_len = len(score_str) + 2  # digits + space + icon char

        # Wrap columns
        reason_lines = textwrap.wrap(reason, W_REASON) or [""]
        recomm_lines = textwrap.wrap(recomm, W_RECOMM) or [""]
        if mdn_link:
            recomm_lines.append(f"{DIM}https://developer.mozilla.org{mdn_link}{RESET}")
        test_lines = textwrap.wrap(display, W_TEST) or [""]

        n_rows = max(len(test_lines), len(reason_lines), len(recomm_lines))
        for i in range(n_rows):
            t  = test_lines[i]   if i < len(test_lines)   else ""
            r  = reason_lines[i] if i < len(reason_lines) else ""
            c  = recomm_lines[i] if i < len(recomm_lines) else ""
            sc = score_col if i == 0 else ""
            sc_pad = " " * (W_SCORE - sc_plain_len) if i == 0 else " " * W_SCORE
            print(f"  {t:<{W_TEST}} {sc}{sc_pad} {r:<{W_REASON}} {c}")
    print()


def print_result(idx: int, total: int, name: str, url: str, result: dict, tests: dict):
    grade = result.get("grade", "N/A")
    score = result.get("score", "N/A")
    error = result.get("error")
    details_url = result.get("details_url", f"https://developer.mozilla.org/en-US/observatory/analyze?host={url}")
    failed = result.get("tests_failed", "?")
    passed = result.get("tests_passed", "?")
    total_tests = result.get("tests_quantity", "?")
    col = grade_color(grade)

    print(f"{BOLD}[{idx}/{total}] {name}{RESET}")
    print(f"  URL       : {url}")

    if error:
        print(f"  {RED}Error     : {error}{RESET}")
        msg = result.get("message", "")
        if msg:
            print(f"  {DIM}Detail    : {msg}{RESET}")
        print(f"  Recomm.   : Verify the hostname is publicly reachable and try again.")
    else:
        print(f"  Grade     : {col}{BOLD}{grade}{RESET}  |  Score: {score}/100")
        print(f"  Tests     : {passed} passed, {failed} failed of {total_tests}")
        print(f"  Report    : {DIM}{details_url}{RESET}")

        if tests:
            _print_test_table(tests)

    print(f"  {DIM}{'-'*60}{RESET}")


# ---------------------------------------------------------------------------
# Result building
# ---------------------------------------------------------------------------

def build_result_row(institution: dict, result: dict, tests: dict) -> dict:
    """Flatten scan result into a CSV-ready dict."""
    grade = result.get("grade", "")
    score = result.get("score", "")
    error = result.get("error", "")
    scanned_at = result.get("scanned_at", "")
    details_url = result.get("details_url", "")
    tests_failed = result.get("tests_failed", "")
    tests_passed = result.get("tests_passed", "")
    tests_quantity = result.get("tests_quantity", "")

    # Collect failing test names
    failing_tests = ""
    if tests:
        failures = [k for k, v in tests.items() if not v.get("pass", True)]
        failing_tests = "; ".join(failures)

    row = {
        "ID":             institution.get("ID", ""),
        "Name":           institution.get("Name", ""),
        "Category":       institution.get("Category", ""),
        "NUTS2":          institution.get("NUTS2", ""),
        "NUTS2_Label":    institution.get("NUTS2_Label", ""),
        "url":            institution.get("url", ""),
        "grade":          grade,
        "score":          score,
        "tests_passed":   tests_passed,
        "tests_failed":   tests_failed,
        "tests_quantity": tests_quantity,
        "failing_tests":  failing_tests,
        "error":          error,
        "details_url":    details_url,
        "scanned_at":     scanned_at,
    }

    # Per-test columns — ordered list first, then any extras the API returned
    all_test_keys = list(KNOWN_TESTS) + [k for k in (tests or {}) if k not in KNOWN_TESTS]
    for test_key in all_test_keys:
        col_prefix = test_key.replace("-", "_")
        tdata  = (tests or {}).get(test_key, {})
        passed = tdata.get("pass", None)
        reason, recomm = _resolve_test(tdata) if tdata else ("", "")
        row[f"{col_prefix}_score"]          = tdata.get("score_modifier", "") if tdata else ""
        row[f"{col_prefix}_pass"]           = ("Pass" if passed else "Fail") if passed is not None else ""
        row[f"{col_prefix}_reason"]         = reason
        row[f"{col_prefix}_recommendation"] = "" if passed else recomm

    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print_header()

    RESULTS_DIR.mkdir(exist_ok=True)

    csv_files = find_source_csvs()
    if not csv_files:
        sys.exit(1)

    all_rows = []

    for csv_path in csv_files:
        print(f"{BOLD}Source file: {csv_path}{RESET}\n")
        institutions = read_institutions(csv_path)
        total = len(institutions)
        print(f"  Found {total} institutions to scan.\n")

        for idx, inst in enumerate(institutions, 1):
            url = inst.get("url", "").strip()
            name = inst.get("Name", url)

            if not url:
                print(f"[{idx}/{total}] {name} - SKIPPED (no URL)\n")
                continue

            # Strip any scheme the CSV might include
            hostname = url.replace("https://", "").replace("http://", "").rstrip("/")

            result = scan_host(hostname)
            tests = {}
            if not result.get("error"):
                tests = get_analyze_data(hostname)

            print_result(idx, total, name, hostname, result, tests)
            all_rows.append(build_result_row(inst, result, tests))

            if idx < total:
                time.sleep(SCAN_COOLDOWN)

    # Write results CSV and XLSX
    if all_rows:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fieldnames = list(all_rows[0].keys())

        out_path = RESULTS_DIR / f"observatory_scan_{timestamp}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_rows)

        xlsx_path = RESULTS_DIR / f"observatory_scan_{timestamp}.xlsx"
        write_xlsx(all_rows, fieldnames, xlsx_path)

        print(f"\n{GREEN}{BOLD}Results saved to: {out_path}{RESET}")
        print(f"{GREEN}{BOLD}Results saved to: {xlsx_path}{RESET}")
        print(f"Total institutions scanned: {len(all_rows)}")

        # Summary stats
        graded = [r for r in all_rows if r["grade"]]
        errors = [r for r in all_rows if r["error"]]
        if graded:
            grades = [r["grade"] for r in graded]
            from collections import Counter
            dist = Counter(grades)
            print(f"\nGrade distribution:")
            for g in ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]:
                if g in dist:
                    col = grade_color(g)
                    print(f"  {col}{g:4s}{RESET}  {dist[g]} institution(s)")
        if errors:
            print(f"\n{YELLOW}Scan errors: {len(errors)}{RESET}")
            for r in errors:
                print(f"  {r['Name']}: {r['error']}")

    print(f"\n{BOLD}Done.{RESET}\n")


if __name__ == "__main__":
    main()
